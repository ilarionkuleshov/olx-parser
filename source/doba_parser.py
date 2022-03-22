import time, json

from openpyxl import Workbook, load_workbook

import requests
from bs4 import BeautifulSoup


class Config:

	doba_url = 'https://doba.ua/kiev/'
	pages = 100
	filename = 'doba.xlsx'


def urls_to_file(doba_url, pages):
	urls_list = []

	for page_num in range(1, pages+1):
		page_url = f"{doba_url}page-{page_num}"

		print(page_num)

		page_response = requests.get(page_url).text
		page_soup = BeautifulSoup(page_response, 'lxml')

		wrappers = page_soup.find_all('div', class_='element-item f f-fdc f-jcsb')

		for wrapper in wrappers:
			picture = wrapper.find('div', class_='element-picture')

			a = picture.find('a', href=True)

			url = f"https://doba.ua/ukr{a['href']}"
			#print(url)

			#if not url in urls_list:
			urls_list.append(url)

	urls_list = list(dict.fromkeys(urls_list))

	with open('doba_urls.json', 'w') as file:
		json.dump({'urls': urls_list}, file)


def get_urls():
	with open('doba_urls.json', 'r') as file:
		urls_list = json.load(file)['urls']

	return urls_list


def values_to_table(filename, values, succes_counter):
	if succes_counter == 0:
		workbook = Workbook()
	else:
		workbook = load_workbook(filename)

	worksheet = workbook.active
	worksheet.append(values)

	workbook.save(filename)


def parse_url(url, is_error=False):
	try:
		url_response = requests.get(url).text
		url_soup = BeautifulSoup(url_response, 'lxml')

		header_text = url_soup.find('h1', class_='d-none d-md-block max_633').text
		author = url_soup.find('div', class_='element-id__user_name').text
		apartment_id = url.split('/')[-1].split('.')[0]

		price_info = url_soup.find('span', class_='element-id__price').text.split(' ')
		price = price_info[0]
		currency = price_info[1].split('.')[0]

		phone_elements = url_soup.find_all('div', class_='element-id__phone_number_item')
		phones = ''

		for phone in phone_elements:
			phones += phone.text
			phones += ','

		phones = phones[:-1]

		values = [header_text, author, phones, price, currency, apartment_id, url]

	except:
		if not is_error:
			values = parse_url(url, True)
		else:
			raise Exception('Ошибка при считывании страницы. Пропуск...')

	return values


if __name__ == '__main__':
	config = Config()

	urls_to_file(config.doba_url, config.pages)
	urls = get_urls()

	urls_n = len(urls)
	succes_counter = 0
	pass_counter = 0

	values_to_table(config.filename, ['Название', 'Автор', 'Телефон(ы)', 'Цена', 'Валюта', 'ID объявления', 'Ссылка'], succes_counter)
	succes_counter += 1

	for url in urls:
		try:
			values = parse_url(url)
			values_to_table(config.filename, values, succes_counter)

			succes_counter += 1

		except Exception as e:
			pass_counter += 1
			print(e)

		print(f"Выполнено: {succes_counter-1}/{urls_n}; пропущено: {pass_counter}")

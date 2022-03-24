import time, json

from openpyxl import Workbook, load_workbook

import requests
from bs4 import BeautifulSoup

from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By


class Config:

	olx_url = 'https://www.olx.ua/elektronika/telefony-i-aksesuary/'
	pages = 25
	filename = 'olx.xlsx'

	path_to_web_driver = './chromedriver/chromedriver.exe'
	proxy = '' # proxy address
	proxy_type = '' # http or socks
	proxy_auth = '' # user and password


class WebDriver:

	def __init__(self, exe_path, proxy, proxy_type, proxy_auth):
		self.exe_path = exe_path
		self.proxy = proxy
		self.proxy_type = proxy_type
		self.proxy_auth = proxy_auth

		self.create()

		self.errors = 0

	def create(self):
		self.driver = Chrome(
			executable_path=self.exe_path,
			service_args=[
				f"--proxy={self.proxy}",
				f"--proxy-type={self.proxy_type}",
				f"--proxy-auth={self.proxy_auth}"
			]
		)
		self.driver.set_window_size(1700, 1000)

	def end(self):
		self.driver.delete_all_cookies()
		self.driver.close()
		self.driver.quit()

	def redefine(self):
		self.end()
		self.create()

	def get(self, *args, **kwargs):
		return self.driver.get(*args, **kwargs)

	def find_element(self, *args, **kwargs):
		return self.driver.find_element(*args, **kwargs)

	def execute_script(self, *args, **kwargs):
		return self.driver.execute_script(*args, **kwargs)


def urls_to_file(olx_url, pages):
	urls_list = []

	for page_num in range(1, pages+1):
		page_url = f"{olx_url}?page={page_num}"

		page_response = requests.get(page_url).text
		page_soup = BeautifulSoup(page_response, 'lxml')

		wrappers = page_soup.find_all('div', class_='offer-wrapper')

		for wrapper in wrappers:
			titles = wrapper.find_all('td', class_='title-cell')

			for title in titles:
				a_tags = title.find_all('a', href=True)

				for a in a_tags:
					url = a['href']
					print(url)

					urls_list.append(url)

	urls_list = list(dict.fromkeys(urls_list))

	with open('olx_urls.json', 'w') as file:
		json.dump({'urls': urls_list}, file)


def get_urls():
	with open('olx_urls.json', 'r') as file:
		urls_list = json.load(file)['urls']

	return urls_list


def values_to_table(filename, values, succes_counter):
	if succes_counter == 0:
		workbook = Workbook()
	else:
		workbook = load_workbook(filename)

	worksheet = workbook.active
	worksheet.append(values)

	workbook.save(filename)


def parse_phones(driver, url):
	try:
		driver.execute_script('window.scrollTo(0, 500)')
		driver.find_element(By.XPATH, "//button[@class='css-1yhx5vv-BaseStyles']").click()
		time.sleep(1)

		phones_list = driver.find_element(By.XPATH, "//ul[@class='css-1478ixo']").text.split('\n')
		phones = ''

		for phone in phones_list:
			phones += phone
			phones += ','

		phones = phones[:-1]

	except:
		if driver.errors >= 4:
			driver.errors = 0
			raise Exception('')

		else:
			driver.errors += 1
			print('Переопределение веб браузера. Подождите...')

			driver.redefine()

			driver.get(url)
			time.sleep(2)

			phones = get_phones(driver, url)

	return phones


def parse_url(driver, url, is_error=False):
	try:
		driver.get(url)
		time.sleep(2)

		header_text = driver.find_element(By.XPATH, "//h1[@class='css-r9zjja-Text eu5v0x0']").text

		author = driver.find_element(By.XPATH, "//h2[@class='css-u8mbra-Text eu5v0x0']").text
		apartment_id = driver.find_element(By.XPATH, "//span[@class='css-9xy3gn-Text eu5v0x0']").text.split(' ')[1]

		price_info = driver.find_element(By.XPATH, "//h3[@class='css-okktvh-Text eu5v0x0']").text.split(' ')
		currency = price_info[-1][:-1]

		price = ''
		for i in range(len(price_info)-1):
			price += price_info[i]

		views = driver.find_element(By.XPATH, "//span[@class='css-1qvxqpo']").text.split(' ')[1]

		phones = parse_phones(driver, url)

		values = [header_text, author, phones, price, currency, views, apartment_id, url]

	except:
		if not is_error:
			values = parse_url(driver, url, True)
		else:
			raise Exception('Ошибка при считывании страницы. Пропуск...')

	return values


if __name__ == '__main__':
	config = Config()

	urls_to_file(config.olx_url, config.pages)

	urls = get_urls()
	driver = WebDriver(config.path_to_web_driver, config.proxy, config.proxy_type, config.proxy_auth)

	urls_n = len(urls)
	succes_counter = 0
	pass_counter = 0

	values_to_table(config.filename, ['Название', 'Автор', 'Телефон(ы)', 'Цена', 'Валюта', 'Просмотры', 'ID объявления', 'Ссылка'], succes_counter)
	succes_counter += 1

	for url in urls:
		try:
			values = parse_url(driver, url)
			values_to_table(config.filename, values, succes_counter)

			succes_counter += 1

		except Exception as e:
			pass_counter += 1
			print(e)

		print(f"Выполнено: {succes_counter-1}/{urls_n}; пропущено: {pass_counter}")

	driver.end()
